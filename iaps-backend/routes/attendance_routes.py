"""attendance_routes.py — Per-session attendance tracking.

Real-world model: like a college register where:
  - CR marks each class as "held" or "cancelled"
  - Students mark themselves present/absent/on-leave (self mode)
  - OR CR takes official roll for a subject (cr mode)
  - System tracks leaves remaining per subject based on per-subject or semester threshold

Collections:
  attendance_sessions        — one per timetable slot per calendar date
  attendance_records         — one per student per session
  attendance_settings        — one per semester (semester-level threshold)
  subject_attendance_config  — one per subject per semester (mode, threshold, periods)

Tracking modes per subject:
  off   — not tracked here (teacher uses IRIS etc.)
  self  — each student marks their own attendance
  cr    — CR marks officially; students are read-only from cr_period_start onward

State machine (giving up control is one-directional):
  off  ↔  self
  off  →  cr   (fresh period, from_date required)
  self →  cr   (CR takes over from today)
  self →  off
  cr   →  off  (export required first; period archived)
  cr   →  self  BLOCKED

Record status values: 'present' | 'absent' | 'leave' | 'college_work'
  'leave' = medical/OD excused — counts as positive attendance.
  'college_work' = official duty excused — counts as positive attendance (same as leave).
  Old records with present: bool are read with backwards-compat fallback.

REST:
  GET   /api/attendance/semester/<id>/settings
  PATCH /api/attendance/semester/<id>/settings
  GET   /api/attendance/semester/<id>/subject-configs
  PATCH /api/attendance/semester/<id>/subject/<subject>/config
  GET   /api/attendance/semester/<id>/summary
  GET   /api/attendance/semester/<id>/sessions
  PATCH /api/attendance/semester/<id>/sessions/<session_id>
  POST  /api/attendance/semester/<id>/mark
  PUT   /api/attendance/semester/<id>/mark/<session_id>
  GET   /api/attendance/semester/<id>/history/<subject>
  GET   /api/attendance/semester/<id>/cr-roll/<session_id>
  POST  /api/attendance/semester/<id>/cr-roll/<session_id>/<student_id>
  POST  /api/attendance/semester/<id>/generate
  GET   /api/attendance/semester/<id>/defaulters
  GET   /api/attendance/semester/<id>/subject/<subject>/export/excel
  GET   /api/attendance/semester/<id>/export/excel
  POST  /api/attendance/semester/<id>/record/<session_id>/attachment
  DELETE /api/attendance/semester/<id>/record/<session_id>/attachment
  GET   /api/attendance/proof/<filename>
"""

import io
import math
import os
import uuid
import logging
from datetime import datetime, date, timedelta, timezone

from flask import Blueprint, request, jsonify, send_file, send_from_directory
from bson import ObjectId

from middleware import token_required, is_member_of_classroom, is_cr_of as _is_cr

attendance_bp = Blueprint('attendance', __name__, url_prefix='/api/attendance')
logger = logging.getLogger(__name__)

VALID_STATUSES = {'present', 'absent', 'leave', 'college_work'}
PROOF_DIR = os.path.join(os.getcwd(), 'uploads', 'attendance_proofs')

# Slot types that do NOT generate attendance sessions
BLOCKED_TYPES = {'Free', 'Lunch', 'Library', 'Break', 'Holiday', 'Cancelled', 'Exam', ''}

# Academic calendar event types that mark a date as no-class
HOLIDAY_EVENT_TYPES = {'Holiday', 'Break', 'No Class'}

# Weekday index → timetable abbreviation (Monday=0)
DAY_MAP = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

# Valid subject-mode transitions. ('cr', 'self') is deliberately absent.
VALID_TRANSITIONS = {
    ('off', 'self'), ('self', 'off'),
    ('off', 'cr'),
    ('self', 'cr'),
    ('cr', 'off'),
}

# Lab/tutorial suffixes — subjects sharing the same base name are grouped together
_LAB_SUFFIXES = (' Lab', ' Laboratory', ' Tutorial', ' Tut', ' Practical')
_LAB_PAREN_RE = None  # lazy compiled below


def _base_subject(name):
    """
    Strip lab/tutorial suffix to get the base subject code.
    Handles both 'IT251 Lab' and 'IT251(lab)' / 'IT251 (Lab)' → 'IT251'.
    """
    import re as _re
    n = name.strip()
    # Parenthetical: "IT251(lab)", "IT251 (Lab)", "CS101(tutorial)" etc.
    paren = _re.sub(
        r'\s*\(\s*(lab|laboratory|tutorial|tut|practical)\s*\)\s*$',
        '', n, flags=_re.IGNORECASE
    ).strip()
    if paren != n:
        return paren
    # Suffix: "IT251 Lab", "CS101 Tutorial" etc.
    for suffix in sorted(_LAB_SUFFIXES, key=len, reverse=True):
        if n.lower().endswith(suffix.lower()):
            return n[:len(n) - len(suffix)].strip()
    return n


def _zone(pct):
    if pct >= 90:
        return 'green'
    if pct >= 80:
        return 'yellow'
    if pct >= 75:
        return 'orange'
    return 'red'


# ── Helpers ──────────────────────────────────────────────────────────────────


def _check_access(db, semester_id, user_id):
    """Return (semester, classroom, is_cr) or (None, None, False)."""
    try:
        semester = db.semesters.find_one({'_id': ObjectId(semester_id)})
    except Exception:
        return None, None, False
    if not semester:
        return None, None, False
    classroom = db.classrooms.find_one({'_id': ObjectId(semester['classroom_id'])})
    if not classroom or not is_member_of_classroom(classroom, user_id):
        return None, None, False
    return semester, classroom, _is_cr(semester, user_id)


def _get_holiday_dates(db, semester_id):
    """Return a set of YYYY-MM-DD strings that are holidays/no-class."""
    cal = db.academic_calendars.find_one({'semester_id': semester_id})
    holiday_dates = set()
    if not cal:
        return holiday_dates, cal
    for ev in cal.get('events', []):
        if ev.get('type') not in HOLIDAY_EVENT_TYPES:
            continue
        ev_start = ev.get('date', '')
        ev_end = ev.get('end_date', ev_start)
        try:
            d = date.fromisoformat(ev_start)
            end_d = date.fromisoformat(ev_end)
            while d <= end_d:
                holiday_dates.add(d.isoformat())
                d += timedelta(days=1)
        except ValueError:
            pass
    return holiday_dates, cal


def _get_override_map(db, semester_id, from_str, to_str):
    """Return {date_str: {slot: 'cancel'|'modify', '__day__': 'cancel'}} from overrides."""
    overrides = list(db.timetable_overrides.find({
        'semester_id': semester_id,
        'date': {'$gte': from_str, '$lte': to_str},
    }))
    ov_map = {}
    for ov in overrides:
        ov_date = ov['date']
        ov_slot = ov.get('slot', '')
        ov_action = ov.get('action', '')
        ov_scope = ov.get('scope', 'slot')
        if ov_date not in ov_map:
            ov_map[ov_date] = {}
        if ov_scope == 'day' and ov_action == 'cancel':
            ov_map[ov_date]['__day__'] = 'cancel'
        elif ov_slot:
            ov_map[ov_date][ov_slot] = ov_action
    return ov_map


def _rec_status(rec):
    """Read status from a record — handles old (present: bool) and new (status: str) schemas.
    No record = absent by default (there is no undefined state).
    """
    if rec is None:
        return 'absent'
    if 'status' in rec:
        return rec['status']
    return 'present' if rec.get('present') else 'absent'


def _get_settings(db, semester_id):
    return db.attendance_settings.find_one({'semester_id': semester_id}) or {}


def _serialize_settings(s):
    if not s:
        return {'threshold': 75.0, 'threshold_log': [], 'version': 0}
    return {
        'threshold': s.get('threshold', 75.0),
        'threshold_log': s.get('threshold_log', []),
        'version': s.get('version', 0),
    }


def _get_subject_config(db, semester_id, subject):
    """Return subject_attendance_config doc, or {} if none."""
    return db.subject_attendance_config.find_one({
        'semester_id': semester_id, 'subject': subject,
    }) or {}


def _subject_mode(db, semester_id, subject, settings=None):
    """
    Tracking mode for a subject. Prefers subject_attendance_config;
    falls back to cr_roll_subjects in settings for backwards compat.
    """
    cfg = db.subject_attendance_config.find_one({
        'semester_id': semester_id, 'subject': subject,
    })
    if cfg:
        return cfg.get('tracking_mode', 'self')
    if settings is None:
        settings = db.attendance_settings.find_one({'semester_id': semester_id}) or {}
    if subject in settings.get('cr_roll_subjects', []):
        return 'cr'
    return 'self'


def _is_student_editable(cfg, session_date_str, settings=None, subject=None):
    """
    Returns True if a student can mark/edit a session on session_date_str.
    cfg is the subject_attendance_config doc (or {}).
    Falls back to settings.cr_roll_subjects if no cfg.
    """
    if not cfg:
        if settings and subject:
            return subject not in settings.get('cr_roll_subjects', [])
        return True

    # Hard lock: CR explicitly locked student marking for this subject
    if cfg.get('locked'):
        return False

    mode = cfg.get('tracking_mode', 'self')
    if mode == 'off':
        return False
    if mode == 'self':
        return True
    # cr mode: locked from cr_period_start onwards
    period_start = cfg.get('cr_period_start')
    if not period_start:
        return False
    return session_date_str < period_start


def _serialize_subject_config(cfg, semester_threshold):
    if not cfg:
        return None
    return {
        'tracking_mode': cfg.get('tracking_mode', 'self'),
        'required_pct': cfg.get('required_pct'),
        'cr_period_start': cfg.get('cr_period_start'),
        'archived_periods': cfg.get('archived_periods', []),
        'last_exported_at': cfg.get('last_exported_at'),
        'locked': cfg.get('locked', False),
        'effective_threshold': (
            cfg['required_pct'] if cfg.get('required_pct') is not None else semester_threshold
        ),
    }



def _check_and_notify_threshold(db, socketio_ref, semester_id, student_id, subject, old_pct, new_pct):
    """Emit a socket notification if attendance drops into a worse zone."""
    zone_rank = {'green': 3, 'yellow': 2, 'orange': 1, 'red': 0}
    if zone_rank[_zone(new_pct)] < zone_rank[_zone(old_pct)]:
        try:
            from routes.chat_routes import _connected_users
            for sid, ud in list(_connected_users.items()):
                if ud.get('user_id') == student_id:
                    socketio_ref.emit('attendance_warning', {
                        'subject': subject,
                        'percentage': new_pct,
                        'zone': _zone(new_pct),
                        'semester_id': semester_id,
                    }, to=sid)
        except Exception:
            pass


def generate_sessions(db, semester_id, classroom_id, from_str, to_str):
    """
    Generate attendance_sessions from timetable for dates from_str..to_str.
    Skips dates that already have sessions. Applies calendar holidays and
    timetable overrides so sessions reflect real-world class schedule.
    Called lazily on summary fetch and explicitly on timetable save.
    """
    tt = db.timetables.find_one({'semester_id': semester_id})
    if not tt:
        return

    grid = tt.get('grid', {})
    days = tt.get('days', [])
    time_slots = tt.get('time_slots', [])
    if not days or not time_slots:
        return

    holiday_dates, _ = _get_holiday_dates(db, semester_id)
    ov_map = _get_override_map(db, semester_id, from_str, to_str)

    existing = set()
    for sess in db.attendance_sessions.find(
        {'semester_id': semester_id, 'date': {'$gte': from_str, '$lte': to_str}},
        {'date': 1, 'slot': 1},
    ):
        existing.add((sess['date'], sess['slot']))

    now = datetime.now(timezone.utc)
    to_insert = []

    try:
        cur = date.fromisoformat(from_str)
        end = date.fromisoformat(to_str)
    except ValueError:
        return

    while cur <= end:
        date_str = cur.isoformat()
        day_abbr = DAY_MAP[cur.weekday()]

        if date_str not in holiday_dates and day_abbr in days:
            day_grid = grid.get(day_abbr, {})
            day_ov = ov_map.get(date_str, {})
            day_cancelled = day_ov.get('__day__') == 'cancel'

            for slot in time_slots:
                if (date_str, slot) in existing:
                    continue

                cell = day_grid.get(slot, {})
                cell_type = cell.get('type', 'Free')
                subject = (cell.get('subject') or '').strip()

                if not subject or cell_type in BLOCKED_TYPES:
                    continue

                if day_cancelled or day_ov.get(slot) == 'cancel':
                    status = 'cancelled'
                else:
                    status = 'pending'

                to_insert.append({
                    'semester_id': semester_id,
                    'classroom_id': classroom_id,
                    'date': date_str,
                    'day': day_abbr,
                    'slot': slot,
                    'subject': subject,
                    'type': cell_type,
                    'status': status,
                    'marked_by_cr_id': None,
                    'marked_at': None,
                    'created_at': now,
                })

        cur += timedelta(days=1)

    if to_insert:
        db.attendance_sessions.insert_many(to_insert)


def _project_future(db, semester_id, today_str, sem_end_str):
    """Count remaining future sessions per subject by walking the timetable forward."""
    tt = db.timetables.find_one({'semester_id': semester_id})
    if not tt or not sem_end_str:
        return {}

    grid = tt.get('grid', {})
    days = tt.get('days', [])
    time_slots = tt.get('time_slots', [])

    holiday_dates, _ = _get_holiday_dates(db, semester_id)

    try:
        tomorrow = date.fromisoformat(today_str) + timedelta(days=1)
        sem_end = date.fromisoformat(sem_end_str)
    except ValueError:
        return {}

    future = {}
    cur = tomorrow
    while cur <= sem_end:
        date_str = cur.isoformat()
        day_abbr = DAY_MAP[cur.weekday()]
        if date_str not in holiday_dates and day_abbr in days:
            day_grid = grid.get(day_abbr, {})
            for slot in time_slots:
                cell = day_grid.get(slot, {})
                cell_type = cell.get('type', 'Free')
                subject = (cell.get('subject') or '').strip()
                if subject and cell_type not in BLOCKED_TYPES:
                    future[subject] = future.get(subject, 0) + 1
        cur += timedelta(days=1)

    return future


def _calc_subject_stats(attended, total, leaves_count, remaining, threshold_pct):
    """
    Returns (pct, leaves_left, must_attend, recoverable) for a subject.
    attended     = sessions marked present (raw count, not including leave)
    total        = all past sessions (happened + pending, excluding cancelled)
    leaves_count = sessions marked 'leave' (medical — counts as positive attendance)
    remaining    = future timetable sessions until semester end
    threshold_pct = e.g. 75.0

    Medical leave counts toward positive attendance: attended_eff = attended + leaves_count.

    Formulas (IRIS-style, current-state unbounded):
      pct         = attended_eff / total * 100
      leaves_left = floor((attended_eff - θ*total) / θ)    → max consecutive classes you can skip right now
      must_attend = ceil((θ*total - attended_eff) / (1-θ)) → min consecutive classes needed to recover
      recoverable = must_attend_raw <= remaining  (False = can't reach threshold even attending everything)

    Both leaves_left and must_attend are capped at `remaining` for display.
    """
    T = threshold_pct / 100.0
    attended_eff = attended + leaves_count   # present + medical leave both count
    pct = round(attended_eff / total * 100, 1) if total > 0 else 0.0

    if T <= 0:
        leaves_left = remaining
        raw_must = 0
    elif T >= 1.0:
        # 100% threshold: any absence means unrecoverable
        leaves_left = 0
        raw_must = remaining if attended_eff < total else 0
    else:
        # IRIS formula: current-state projection (not bounded to semester end)
        leaves_left = max(0, math.floor((attended_eff - T * total) / T))
        raw_must = max(0, math.ceil((T * total - attended_eff) / (1.0 - T)))

    recoverable = raw_must <= remaining
    must_attend = min(raw_must, remaining)
    leaves_left = min(leaves_left, remaining)

    return pct, leaves_left, must_attend, recoverable


# ── Settings ─────────────────────────────────────────────────────────────────


@attendance_bp.route('/semester/<semester_id>/settings', methods=['GET'])
@token_required
def get_settings(semester_id):
    from database import get_db
    try:
        user_id = request.user['user_id']
        db = get_db()
        semester, _, _ = _check_access(db, semester_id, user_id)
        if not semester:
            return jsonify({'error': 'Access denied'}), 403
        s = _get_settings(db, semester_id)
        return jsonify({'settings': _serialize_settings(s)}), 200
    except Exception as e:
        logger.error(f"Get attendance settings error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


@attendance_bp.route('/semester/<semester_id>/settings', methods=['PATCH'])
@token_required
def update_settings(semester_id):
    """CR updates semester-level threshold (optimistic-locked)."""
    from database import get_db
    try:
        user_id = request.user['user_id']
        db = get_db()
        semester, classroom, is_cr = _check_access(db, semester_id, user_id)
        if not semester or not is_cr:
            return jsonify({'error': 'CR access required'}), 403

        data = request.get_json() or {}
        now = datetime.now(timezone.utc)
        classroom_id = str(semester['classroom_id'])

        existing = db.attendance_settings.find_one({'semester_id': semester_id}) or {}
        update = {}

        if 'threshold' in data:
            client_version = data.get('version', 0)
            db_version = existing.get('version', 0)
            if client_version != db_version:
                return jsonify({'error': 'Conflict — settings were updated by another CR. Refresh and try again.'}), 409

            new_threshold = float(data['threshold'])
            if not (50.0 <= new_threshold <= 100.0):
                return jsonify({'error': 'Threshold must be between 50 and 100'}), 400

            old_threshold = existing.get('threshold', 75.0)
            update['threshold'] = new_threshold
            update['version'] = db_version + 1
            if new_threshold != old_threshold:
                update['$push_threshold_log'] = {
                    'from': old_threshold,
                    'to': new_threshold,
                    'changed_by': user_id,
                    'changed_by_name': request.user.get('username', ''),
                    'changed_at': now.isoformat(),
                }

        if not update:
            return jsonify({'error': 'Nothing to update'}), 400

        update['updated_at'] = now
        set_fields = {k: v for k, v in update.items() if k != '$push_threshold_log'}
        mongo_update = {'$set': set_fields}
        if '$push_threshold_log' in update:
            mongo_update['$push'] = {'threshold_log': update['$push_threshold_log']}

        if existing:
            db.attendance_settings.update_one({'semester_id': semester_id}, mongo_update)
        else:
            new_doc = {
                'semester_id': semester_id,
                'classroom_id': classroom_id,
                'threshold': update.get('threshold', 75.0),
                'version': update.get('version', 1),
                'threshold_log': [update['$push_threshold_log']] if '$push_threshold_log' in update else [],
                'created_at': now,
                'updated_at': now,
            }
            db.attendance_settings.insert_one(new_doc)

        s = _get_settings(db, semester_id)
        return jsonify({'settings': _serialize_settings(s)}), 200
    except Exception as e:
        logger.error(f"Update attendance settings error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


# ── Subject configs ───────────────────────────────────────────────────────────


@attendance_bp.route('/semester/<semester_id>/subject-configs', methods=['GET'])
@token_required
def get_subject_configs(semester_id):
    """All subject configs for this semester. Returns a map of subject → config."""
    from database import get_db
    try:
        user_id = request.user['user_id']
        db = get_db()
        semester, _, _ = _check_access(db, semester_id, user_id)
        if not semester:
            return jsonify({'error': 'Access denied'}), 403

        settings = _get_settings(db, semester_id)
        threshold = settings.get('threshold', 75.0)

        configs = list(db.subject_attendance_config.find({'semester_id': semester_id}))
        result = {}
        for cfg in configs:
            result[cfg['subject']] = _serialize_subject_config(cfg, threshold)

        # Also surface subjects that only exist in cr_roll_subjects (backwards compat)
        for subj in settings.get('cr_roll_subjects', []):
            if subj not in result:
                result[subj] = _serialize_subject_config(
                    {'tracking_mode': 'cr', 'subject': subj}, threshold
                )

        return jsonify({'configs': result, 'semester_threshold': threshold}), 200
    except Exception as e:
        logger.error(f"Get subject configs error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


@attendance_bp.route('/semester/<semester_id>/subject/<path:subject>/config', methods=['PATCH'])
@token_required
def update_subject_config(semester_id, subject):
    """
    CR updates a subject's tracking mode or required_pct.

    Mode transitions enforced — cr → self is blocked.
    cr → off requires export to have happened at least once (last_exported_at set).
    off → cr requires cr_period_start in request body.
    """
    from database import get_db
    try:
        user_id = request.user['user_id']
        db = get_db()
        semester, _, is_cr = _check_access(db, semester_id, user_id)
        if not semester or not is_cr:
            return jsonify({'error': 'CR access required'}), 403

        data = request.get_json() or {}
        now = datetime.now(timezone.utc)
        classroom_id = str(semester['classroom_id'])

        existing = db.subject_attendance_config.find_one({
            'semester_id': semester_id, 'subject': subject,
        })
        current_mode = existing.get('tracking_mode', 'self') if existing else 'self'

        # ── Mode transition ──
        if 'tracking_mode' in data:
            new_mode = data['tracking_mode']
            if new_mode not in ('off', 'self', 'cr'):
                return jsonify({'error': 'tracking_mode must be off, self, or cr'}), 400

            if new_mode != current_mode:
                if (current_mode, new_mode) not in VALID_TRANSITIONS:
                    return jsonify({
                        'error': f"Cannot switch from '{current_mode}' to '{new_mode}'. "
                                 f"Once CR takes official control, control cannot be handed to students. "
                                 f"Switch to 'off' first (export required), then to 'cr' again from a new date."
                    }), 400

                if current_mode == 'cr' and new_mode == 'off':
                    # Must have exported at least once before archiving
                    if existing and not existing.get('last_exported_at'):
                        return jsonify({
                            'error': 'Export attendance before archiving. Download the Excel sheet first.'
                        }), 400
                    # Archive the current CR period
                    period_start = (existing or {}).get('cr_period_start')
                    archived_entry = {
                        'from': period_start,
                        'to': date.today().isoformat(),
                        'exported_at': (existing or {}).get('last_exported_at'),
                    }
                    db.subject_attendance_config.update_one(
                        {'semester_id': semester_id, 'subject': subject},
                        {
                            '$set': {
                                'tracking_mode': 'off',
                                'cr_period_start': None,
                                'updated_at': now,
                            },
                            '$push': {'archived_periods': archived_entry},
                        },
                        upsert=True,
                    )
                    return jsonify({'ok': True, 'tracking_mode': 'off'}), 200

                if new_mode == 'cr':
                    cr_period_start = data.get('cr_period_start')
                    if not cr_period_start:
                        cr_period_start = date.today().isoformat()

                    # Validate: must be after last archived period end
                    archived = (existing or {}).get('archived_periods', [])
                    if archived:
                        last_end = max(p.get('to', '') for p in archived)
                        if cr_period_start <= last_end:
                            return jsonify({
                                'error': f"cr_period_start must be after last archived period end ({last_end})"
                            }), 400

                    db.subject_attendance_config.update_one(
                        {'semester_id': semester_id, 'subject': subject},
                        {'$set': {
                            'semester_id': semester_id,
                            'classroom_id': classroom_id,
                            'subject': subject,
                            'tracking_mode': 'cr',
                            'cr_period_start': cr_period_start,
                            'updated_at': now,
                        }, '$setOnInsert': {'created_at': now, 'archived_periods': []}},
                        upsert=True,
                    )
                    return jsonify({'ok': True, 'tracking_mode': 'cr', 'cr_period_start': cr_period_start}), 200

                # off/self transition
                db.subject_attendance_config.update_one(
                    {'semester_id': semester_id, 'subject': subject},
                    {'$set': {
                        'semester_id': semester_id,
                        'classroom_id': classroom_id,
                        'subject': subject,
                        'tracking_mode': new_mode,
                        'updated_at': now,
                    }, '$setOnInsert': {'created_at': now, 'archived_periods': []}},
                    upsert=True,
                )

        # ── Lock / unlock student marking ──
        if 'locked' in data:
            db.subject_attendance_config.update_one(
                {'semester_id': semester_id, 'subject': subject},
                {'$set': {
                    'semester_id': semester_id,
                    'classroom_id': classroom_id,
                    'subject': subject,
                    'locked': bool(data['locked']),
                    'updated_at': now,
                }, '$setOnInsert': {'created_at': now, 'archived_periods': [], 'tracking_mode': 'self'}},
                upsert=True,
            )

        # ── Per-subject threshold ──
        if 'required_pct' in data:
            val = data['required_pct']
            if val is not None:
                val = float(val)
                if not (50.0 <= val <= 100.0):
                    return jsonify({'error': 'required_pct must be between 50 and 100'}), 400
            db.subject_attendance_config.update_one(
                {'semester_id': semester_id, 'subject': subject},
                {'$set': {
                    'semester_id': semester_id,
                    'classroom_id': classroom_id,
                    'subject': subject,
                    'required_pct': val,
                    'updated_at': now,
                }, '$setOnInsert': {'created_at': now, 'archived_periods': [], 'tracking_mode': 'self'}},
                upsert=True,
            )

        cfg = _get_subject_config(db, semester_id, subject)
        settings = _get_settings(db, semester_id)
        return jsonify({
            'ok': True,
            'config': _serialize_subject_config(cfg, settings.get('threshold', 75.0)),
        }), 200
    except Exception as e:
        logger.error(f"Update subject config error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


# ── Session generation ────────────────────────────────────────────────────────


@attendance_bp.route('/semester/<semester_id>/generate', methods=['POST'])
@token_required
def trigger_generate(semester_id):
    """Called when timetable is saved to regenerate pending sessions."""
    from database import get_db
    try:
        user_id = request.user['user_id']
        db = get_db()
        semester, classroom, is_cr = _check_access(db, semester_id, user_id)
        if not semester or not is_cr:
            return jsonify({'error': 'CR access required'}), 403

        classroom_id = str(semester['classroom_id'])
        _, cal = _get_holiday_dates(db, semester_id)
        today_str = date.today().isoformat()
        sem_start = (cal or {}).get('semester_start', today_str)
        sem_end = (cal or {}).get('semester_end', today_str)
        to_str = min(today_str, sem_end)

        generate_sessions(db, semester_id, classroom_id, sem_start, to_str)
        return jsonify({'ok': True}), 200
    except Exception as e:
        logger.error(f"Generate sessions error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


# ── Summary ───────────────────────────────────────────────────────────────────


@attendance_bp.route('/semester/<semester_id>/summary', methods=['GET'])
@token_required
def get_summary(semester_id):
    """
    Per-subject attendance summary for the calling student.
    Uses per-subject threshold if set, else semester default.
    Medical leave counts as positive attendance (attended_eff = present + leave).
    leaves_left = how many of the remaining sessions you can skip.
    recoverable = whether reaching threshold is still possible this semester.
    """
    from database import get_db
    try:
        user_id = request.user['user_id']
        db = get_db()
        semester, classroom, is_cr = _check_access(db, semester_id, user_id)
        if not semester:
            return jsonify({'error': 'Access denied'}), 403

        classroom_id = str(semester['classroom_id'])
        today_str = date.today().isoformat()

        _, cal = _get_holiday_dates(db, semester_id)
        sem_start = (cal or {}).get('semester_start') if cal else None
        sem_end = (cal or {}).get('semester_end') if cal else None

        if sem_start:
            to_str = min(today_str, sem_end) if sem_end else today_str
            generate_sessions(db, semester_id, classroom_id, sem_start, to_str)

        settings = _get_settings(db, semester_id)
        sem_threshold = settings.get('threshold', 75.0)

        # All held + student-marked sessions up to today
        # (pending = class happened but not yet officially confirmed by CR;
        #  students can already mark themselves on pending sessions)
        happened = list(db.attendance_sessions.find({
            'semester_id': semester_id,
            'status': {'$in': ['happened', 'pending']},
            'date': {'$lte': today_str},
        }))

        # Student's attendance records
        records = {
            str(r['session_id']): r
            for r in db.attendance_records.find({
                'semester_id': semester_id,
                'student_id': user_id,
            })
        }

        # Per-subject: {subject: {attended, total, leaves_count, cw_count, last_date}}
        week_ago = (date.today() - timedelta(days=7)).isoformat()
        subj_data = {}
        past_subj_data = {}  # same but only sessions older than 7 days (for pct_delta)
        for sess in happened:
            subj = sess['subject']
            if subj not in subj_data:
                subj_data[subj] = {'attended': 0, 'total': 0, 'leaves_count': 0, 'cw_count': 0, 'last_date': None}
            subj_data[subj]['total'] += 1
            rec = records.get(str(sess['_id']))
            status = _rec_status(rec)
            if status == 'present':
                subj_data[subj]['attended'] += 1
            elif status == 'leave':
                subj_data[subj]['leaves_count'] += 1
            elif status == 'college_work':
                subj_data[subj]['cw_count'] += 1
            if subj_data[subj]['last_date'] is None or sess['date'] > subj_data[subj]['last_date']:
                subj_data[subj]['last_date'] = sess['date']
            # Past snapshot (7+ days ago)
            if sess['date'] <= week_ago:
                if subj not in past_subj_data:
                    past_subj_data[subj] = {'attended': 0, 'total': 0, 'leaves_count': 0, 'cw_count': 0}
                past_subj_data[subj]['total'] += 1
                if status == 'present':
                    past_subj_data[subj]['attended'] += 1
                elif status == 'leave':
                    past_subj_data[subj]['leaves_count'] += 1
                elif status == 'college_work':
                    past_subj_data[subj]['cw_count'] += 1

        future = _project_future(db, semester_id, today_str, sem_end) if sem_end else {}

        # All subject configs
        all_configs = {
            c['subject']: c
            for c in db.subject_attendance_config.find({'semester_id': semester_id})
        }

        all_subjects = sorted(set(list(subj_data.keys()) + list(future.keys())))

        # Build timetable_name → subject_id map for cross-linking to marks/resources
        subject_id_map = {}
        for s in db.subjects.find({'semester_id': semester_id, 'personal': {'$ne': True}}):
            tt_name = s.get('timetable_name') or s.get('name', '')
            if tt_name:
                subject_id_map[tt_name] = str(s['_id'])

        # Group variants by base name (IT251 Lab → IT251, IT253 Tutorial → IT253, etc.)
        # Base name only appears once; labs/tutorials are merged into their parent.
        variant_to_base = {s: _base_subject(s) for s in all_subjects}
        base_to_variants = {}
        for s in all_subjects:
            base = variant_to_base[s]
            base_to_variants.setdefault(base, []).append(s)

        # Deduplicated base-name list preserving timetable order
        seen_bases = set()
        ordered_bases = []
        for s in all_subjects:
            base = variant_to_base[s]
            if base not in seen_bases:
                seen_bases.add(base)
                ordered_bases.append(base)

        result = []
        for base_name in ordered_bases:
            variants = base_to_variants[base_name]

            # Config / mode from base subject
            cfg = all_configs.get(base_name, {})
            mode = cfg.get('tracking_mode', 'self') if cfg else _subject_mode(db, semester_id, base_name, settings)
            if mode == 'off':
                continue

            # Combine stats across all variants (e.g. IT251 + IT251 Lab)
            combined = {'attended': 0, 'total': 0, 'leaves_count': 0, 'cw_count': 0, 'last_date': None}
            for v in variants:
                d = subj_data.get(v, {'attended': 0, 'total': 0, 'leaves_count': 0, 'cw_count': 0, 'last_date': None})
                combined['attended'] += d['attended']
                combined['total'] += d['total']
                combined['leaves_count'] += d['leaves_count']
                combined['cw_count'] += d.get('cw_count', 0)
                if d['last_date'] and (combined['last_date'] is None or d['last_date'] > combined['last_date']):
                    combined['last_date'] = d['last_date']

            threshold = cfg.get('required_pct') if cfg and cfg.get('required_pct') is not None else sem_threshold
            R = sum(future.get(v, 0) for v in variants)
            excused = combined['leaves_count'] + combined['cw_count']

            pct, leaves_left, must_attend, recoverable = _calc_subject_stats(
                combined['attended'], combined['total'], excused, R, threshold
            )

            # % delta vs 7 days ago
            past_combined = {'attended': 0, 'total': 0, 'leaves_count': 0, 'cw_count': 0}
            for v in variants:
                d = past_subj_data.get(v, {})
                past_combined['attended'] += d.get('attended', 0)
                past_combined['total'] += d.get('total', 0)
                past_combined['leaves_count'] += d.get('leaves_count', 0)
                past_combined['cw_count'] += d.get('cw_count', 0)
            if past_combined['total'] > 0:
                past_eff = past_combined['attended'] + past_combined['leaves_count'] + past_combined.get('cw_count', 0)
                past_pct = round(past_eff / past_combined['total'] * 100, 1)
                pct_delta = round(pct - past_pct, 1)
            else:
                pct_delta = None

            result.append({
                'subject': base_name,
                'variants': variants,
                'subject_id': subject_id_map.get(base_name),
                'attended': combined['attended'],
                'total': combined['total'],
                'leaves_count': combined['leaves_count'],
                'cw_count': combined['cw_count'],
                'percentage': pct,
                'remaining': R,
                'leaves_left': leaves_left,
                'must_attend': must_attend,
                'recoverable': recoverable,
                'pct_delta': pct_delta,
                'zone': _zone(pct) if combined['total'] > 0 else 'green',
                'last_marked_date': combined['last_date'],
                'mode': mode,
                'threshold': threshold,
                'cr_period_start': cfg.get('cr_period_start') if cfg else None,
            })

        return jsonify({
            'subjects': result,
            'threshold': sem_threshold,
            'threshold_log': settings.get('threshold_log', []),
            'settings': _serialize_settings(settings),
        }), 200
    except Exception as e:
        logger.error(f"Get attendance summary error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


# ── Sessions list ─────────────────────────────────────────────────────────────


@attendance_bp.route('/semester/<semester_id>/sessions', methods=['GET'])
@token_required
def list_sessions(semester_id):
    """
    List sessions for a date (default today).
    Students see only happened sessions; CR sees all.
    """
    from database import get_db
    try:
        user_id = request.user['user_id']
        db = get_db()
        semester, classroom, is_cr = _check_access(db, semester_id, user_id)
        if not semester:
            return jsonify({'error': 'Access denied'}), 403

        target_date = request.args.get('date', date.today().isoformat())

        query = {'semester_id': semester_id, 'date': target_date}
        if not is_cr:
            query['status'] = 'happened'

        sessions = list(db.attendance_sessions.find(query).sort('slot', 1))

        records = {}
        if not is_cr:
            for r in db.attendance_records.find({
                'semester_id': semester_id,
                'student_id': user_id,
                'date': target_date,
            }):
                records[str(r['session_id'])] = r

        settings = _get_settings(db, semester_id)

        result = []
        for sess in sessions:
            sess_id = str(sess['_id'])
            cfg = _get_subject_config(db, semester_id, sess['subject'])
            mode = cfg.get('tracking_mode', 'self') if cfg else _subject_mode(db, semester_id, sess['subject'], settings)
            editable = _is_student_editable(cfg or None, target_date, settings, sess['subject'])

            entry = {
                'id': sess_id,
                'date': sess['date'],
                'day': sess['day'],
                'slot': sess['slot'],
                'subject': sess['subject'],
                'type': sess['type'],
                'status': sess['status'],
                'mode': mode,
                'student_editable': editable,
            }
            if not is_cr:
                rec = records.get(sess_id)
                entry['my_record'] = {
                    'status': _rec_status(rec),
                    'marked_by': rec.get('marked_by') if rec else None,
                } if rec else None
            result.append(entry)

        return jsonify({'sessions': result, 'is_cr': is_cr}), 200
    except Exception as e:
        logger.error(f"List sessions error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


# ── CR marks session happened/cancelled ──────────────────────────────────────


@attendance_bp.route('/semester/<semester_id>/sessions/<session_id>', methods=['PATCH'])
@token_required
def mark_session(semester_id, session_id):
    """CR marks a session as happened or cancelled."""
    from database import get_db
    try:
        user_id = request.user['user_id']
        db = get_db()
        semester, classroom, is_cr = _check_access(db, semester_id, user_id)
        if not semester or not is_cr:
            return jsonify({'error': 'CR access required'}), 403

        try:
            sess = db.attendance_sessions.find_one({
                '_id': ObjectId(session_id),
                'semester_id': semester_id,
            })
        except Exception:
            return jsonify({'error': 'Session not found'}), 404
        if not sess:
            return jsonify({'error': 'Session not found'}), 404

        data = request.get_json() or {}
        new_status = data.get('status')
        if new_status not in ('happened', 'cancelled'):
            return jsonify({'error': 'status must be happened or cancelled'}), 400

        now = datetime.now(timezone.utc)

        if new_status == 'cancelled' and sess['status'] == 'happened':
            db.attendance_records.delete_many({'session_id': session_id})

        db.attendance_sessions.update_one(
            {'_id': ObjectId(session_id)},
            {'$set': {
                'status': new_status,
                'marked_by_cr_id': user_id,
                'marked_at': now,
            }},
        )

        if new_status == 'happened':
            try:
                from socketio_instance import socketio
                from routes.chat_routes import _connected_users
                classroom_id = str(semester['classroom_id'])
                for sid, ud in list(_connected_users.items()):
                    if ud.get('classroom_id') == classroom_id or ud.get('user_id'):
                        socketio.emit('attendance_session_happened', {
                            'session_id': session_id,
                            'semester_id': semester_id,
                            'subject': sess['subject'],
                            'slot': sess['slot'],
                            'date': sess['date'],
                        }, to=sid)
            except Exception:
                pass

        return jsonify({'ok': True, 'status': new_status}), 200
    except Exception as e:
        logger.error(f"Mark session error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


# ── Student marks themselves ──────────────────────────────────────────────────


@attendance_bp.route('/semester/<semester_id>/mark', methods=['POST'])
@token_required
def mark_self(semester_id):
    """
    Student marks a session as present, absent, or leave.
    Blocked if subject is in CR mode for that session's date.
    """
    from database import get_db
    try:
        user_id = request.user['user_id']
        db = get_db()
        semester, _, is_cr = _check_access(db, semester_id, user_id)
        if not semester:
            return jsonify({'error': 'Access denied'}), 403

        data = request.get_json() or {}
        session_id = data.get('session_id', '')
        status = data.get('status', 'present')
        if status not in VALID_STATUSES:
            return jsonify({'error': 'Invalid status'}), 400

        try:
            sess = db.attendance_sessions.find_one({
                '_id': ObjectId(session_id),
                'semester_id': semester_id,
                'status': {'$in': ['happened', 'pending']},
            })
        except Exception:
            return jsonify({'error': 'Session not found'}), 404
        if not sess:
            return jsonify({'error': 'Session not found or cancelled'}), 404

        cfg = _get_subject_config(db, semester_id, sess['subject'])
        settings = _get_settings(db, semester_id)
        if not _is_student_editable(cfg or None, sess['date'], settings, sess['subject']):
            return jsonify({'error': 'Attendance is taken by CR for this subject'}), 403

        now = datetime.now(timezone.utc)

        # Old % for threshold notification
        old_records = list(db.attendance_records.find({
            'semester_id': semester_id, 'student_id': user_id, 'subject': sess['subject'],
        }))
        old_N = db.attendance_sessions.count_documents({
            'semester_id': semester_id, 'subject': sess['subject'], 'status': 'happened',
        })
        old_excused = sum(1 for r in old_records if _rec_status(r) in ('leave', 'college_work'))
        old_A = sum(1 for r in old_records if _rec_status(r) == 'present')
        old_pct = round(old_A / max(old_N - old_excused, 1) * 100, 1) if old_N > 0 else 0.0

        db.attendance_records.update_one(
            {'session_id': session_id, 'student_id': user_id},
            {'$set': {
                'session_id': session_id,
                'semester_id': semester_id,
                'classroom_id': str(semester['classroom_id']),
                'student_id': user_id,
                'subject': sess['subject'],
                'date': sess['date'],
                'status': status,
                'marked_by': 'self',
                'updated_at': now,
            }, '$setOnInsert': {'created_at': now}},
            upsert=True,
        )

        new_records = list(db.attendance_records.find({
            'semester_id': semester_id, 'student_id': user_id, 'subject': sess['subject'],
        }))
        new_excused = sum(1 for r in new_records if _rec_status(r) in ('leave', 'college_work'))
        new_A = sum(1 for r in new_records if _rec_status(r) == 'present')
        new_pct = round(new_A / max(old_N - new_excused, 1) * 100, 1) if old_N > 0 else 0.0

        try:
            from socketio_instance import socketio
            _check_and_notify_threshold(db, socketio, semester_id, user_id, sess['subject'], old_pct, new_pct)
        except Exception:
            pass

        return jsonify({'ok': True, 'status': status}), 200
    except Exception as e:
        logger.error(f"Mark self error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


@attendance_bp.route('/semester/<semester_id>/mark/<session_id>', methods=['PUT'])
@token_required
def change_mark(semester_id, session_id):
    """Student changes their mark on any session (no cutoff)."""
    from database import get_db
    try:
        user_id = request.user['user_id']
        db = get_db()
        semester, _, _ = _check_access(db, semester_id, user_id)
        if not semester:
            return jsonify({'error': 'Access denied'}), 403

        data = request.get_json() or {}
        status = data.get('status', 'present')
        if status not in VALID_STATUSES:
            return jsonify({'error': 'Invalid status'}), 400

        try:
            sess = db.attendance_sessions.find_one({
                '_id': ObjectId(session_id),
                'semester_id': semester_id,
                'status': {'$in': ['happened', 'pending']},
            })
        except Exception:
            return jsonify({'error': 'Session not found'}), 404
        if not sess:
            return jsonify({'error': 'Session not found or cancelled'}), 404

        cfg = _get_subject_config(db, semester_id, sess['subject'])
        settings = _get_settings(db, semester_id)
        if not _is_student_editable(cfg or None, sess['date'], settings, sess['subject']):
            return jsonify({'error': 'Attendance is taken by CR for this subject'}), 403

        now = datetime.now(timezone.utc)

        old_records = list(db.attendance_records.find({
            'semester_id': semester_id, 'student_id': user_id, 'subject': sess['subject'],
        }))
        old_N = db.attendance_sessions.count_documents({
            'semester_id': semester_id, 'subject': sess['subject'], 'status': 'happened',
        })
        old_excused = sum(1 for r in old_records if _rec_status(r) in ('leave', 'college_work'))
        old_A = sum(1 for r in old_records if _rec_status(r) == 'present')
        old_pct = round(old_A / max(old_N - old_excused, 1) * 100, 1) if old_N > 0 else 0.0

        db.attendance_records.update_one(
            {'session_id': session_id, 'student_id': user_id},
            {'$set': {
                'session_id': session_id,
                'semester_id': semester_id,
                'classroom_id': str(semester['classroom_id']),
                'student_id': user_id,
                'subject': sess['subject'],
                'date': sess['date'],
                'status': status,
                'marked_by': 'self',
                'updated_at': now,
            }, '$setOnInsert': {'created_at': now}},
            upsert=True,
        )

        new_records = list(db.attendance_records.find({
            'semester_id': semester_id, 'student_id': user_id, 'subject': sess['subject'],
        }))
        new_excused = sum(1 for r in new_records if _rec_status(r) in ('leave', 'college_work'))
        new_A = sum(1 for r in new_records if _rec_status(r) == 'present')
        new_pct = round(new_A / max(old_N - new_excused, 1) * 100, 1) if old_N > 0 else 0.0

        try:
            from socketio_instance import socketio
            _check_and_notify_threshold(db, socketio, semester_id, user_id, sess['subject'], old_pct, new_pct)
        except Exception:
            pass

        return jsonify({'ok': True, 'status': status}), 200
    except Exception as e:
        logger.error(f"Change mark error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


# ── Session history ───────────────────────────────────────────────────────────


@attendance_bp.route('/semester/<semester_id>/history/<path:subject>', methods=['GET'])
@token_required
def get_history(semester_id, subject):
    """Full session history for a subject — all dates (happened + pending)."""
    from database import get_db
    try:
        user_id = request.user['user_id']
        db = get_db()
        semester, _, _ = _check_access(db, semester_id, user_id)
        if not semester:
            return jsonify({'error': 'Access denied'}), 403

        # Include lab/tutorial variants (e.g. viewing 'IT251' also shows 'IT251 Lab')
        base = _base_subject(subject)
        all_sess_subjects = db.attendance_sessions.distinct(
            'subject', {'semester_id': semester_id, 'status': {'$in': ['happened', 'pending']}}
        )
        variants = [s for s in all_sess_subjects if _base_subject(s) == base] or [subject]

        sessions = list(db.attendance_sessions.find({
            'semester_id': semester_id,
            'subject': {'$in': variants},
            'status': {'$in': ['happened', 'pending']},
        }).sort([('date', 1), ('slot', 1)]))

        records = {
            str(r['session_id']): r
            for r in db.attendance_records.find({
                'semester_id': semester_id,
                'student_id': user_id,
                'subject': {'$in': variants},
            })
        }

        cfg = _get_subject_config(db, semester_id, subject)
        settings = _get_settings(db, semester_id)

        result = []
        for sess in sessions:
            sess_id = str(sess['_id'])
            rec = records.get(sess_id)
            editable = _is_student_editable(cfg or None, sess['date'], settings, sess['subject'])
            attachment = None
            if rec and rec.get('attachment'):
                att = rec['attachment']
                attachment = {
                    'original_name': att.get('original_name'),
                    'stored_name': att.get('stored_name'),
                }
            result.append({
                'session_id': sess_id,
                'date': sess['date'],
                'slot': sess['slot'],
                'subject_variant': sess['subject'],
                'type': sess['type'],
                'status': sess['status'],
                'my_status': _rec_status(rec),
                'marked_by': rec.get('marked_by') if rec else None,
                'student_editable': editable,
                'attachment': attachment,
            })

        return jsonify({'history': result, 'subject': subject}), 200
    except Exception as e:
        logger.error(f"Get history error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


# ── CR roll ───────────────────────────────────────────────────────────────────


@attendance_bp.route('/semester/<semester_id>/cr-roll/<session_id>', methods=['GET'])
@token_required
def get_cr_roll(semester_id, session_id):
    """CR fetches the student roster for a session with their current records."""
    from database import get_db
    try:
        user_id = request.user['user_id']
        db = get_db()
        semester, classroom, is_cr = _check_access(db, semester_id, user_id)
        if not semester or not is_cr:
            return jsonify({'error': 'CR access required'}), 403

        try:
            sess = db.attendance_sessions.find_one({
                '_id': ObjectId(session_id),
                'semester_id': semester_id,
            })
        except Exception:
            return jsonify({'error': 'Session not found'}), 404
        if not sess:
            return jsonify({'error': 'Session not found'}), 404

        # CR can always take roll regardless of subject mode

        members = list(db.users.find(
            {'_id': {'$in': classroom.get('members', [])}},
            {'_id': 1, 'username': 1, 'fullName': 1, 'profile_picture': 1, 'roll_number': 1},
        ).sort('roll_number', 1))

        records = {
            r['student_id']: r
            for r in db.attendance_records.find({'session_id': session_id})
        }

        result = []
        for m in members:
            mid = str(m['_id'])
            rec = records.get(mid)
            result.append({
                'user_id': mid,
                'username': m.get('username', ''),
                'full_name': m.get('fullName', ''),
                'roll_number': m.get('roll_number', ''),
                'profile_picture': m.get('profile_picture'),
                'status': _rec_status(rec),  # None = not marked yet
                'marked_at': rec['updated_at'].isoformat() if rec and rec.get('updated_at') else None,
            })

        return jsonify({
            'session': {
                'id': session_id,
                'date': sess['date'],
                'slot': sess['slot'],
                'subject': sess['subject'],
                'status': sess['status'],
            },
            'students': result,
        }), 200
    except Exception as e:
        logger.error(f"Get CR roll error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


@attendance_bp.route('/semester/<semester_id>/subject/<path:subject>/cr-summary', methods=['GET'])
@token_required
def get_cr_subject_summary(semester_id, subject):
    """CR: per-student aggregate attendance for a CR Official subject."""
    from database import get_db
    try:
        user_id = request.user['user_id']
        db = get_db()
        semester, classroom, is_cr = _check_access(db, semester_id, user_id)
        if not semester or not is_cr:
            return jsonify({'error': 'CR access required'}), 403

        base = _base_subject(subject)
        all_subjects = db.attendance_sessions.distinct(
            'subject', {'semester_id': semester_id, 'status': 'happened'}
        )
        variants = [s for s in all_subjects if _base_subject(s) == base] or [subject]

        sessions = list(db.attendance_sessions.find({
            'semester_id': semester_id,
            'subject': {'$in': variants},
            'status': 'happened',
        }))
        total_sessions = len(sessions)
        session_ids = [str(s['_id']) for s in sessions]

        records = list(db.attendance_records.find({
            'semester_id': semester_id,
            'session_id': {'$in': session_ids},
            'subject': {'$in': variants},
        }))

        # Aggregate per student
        agg = {}  # student_id -> {present, absent, leave, college_work}
        for rec in records:
            sid = rec['student_id']
            if sid not in agg:
                agg[sid] = {'present': 0, 'absent': 0, 'leave': 0, 'college_work': 0}
            st = rec.get('status', 'absent')
            if st in agg[sid]:
                agg[sid][st] += 1

        members = list(db.users.find(
            {'_id': {'$in': classroom.get('members', [])}},
            {'_id': 1, 'username': 1, 'fullName': 1, 'profile_picture': 1, 'roll_number': 1},
        ).sort('roll_number', 1))

        cfg = _get_subject_config(db, semester_id, subject)
        settings = _get_settings(db, semester_id)
        threshold = (cfg.get('threshold') if cfg else None) or settings.get('default_threshold', 75)

        result = []
        for m in members:
            mid = str(m['_id'])
            counts = agg.get(mid, {'present': 0, 'absent': 0, 'leave': 0, 'college_work': 0})
            attended = counts['present'] + counts['leave'] + counts['college_work']
            pct = round(attended / total_sessions * 100, 1) if total_sessions > 0 else None
            result.append({
                'user_id': mid,
                'username': m.get('username', ''),
                'full_name': m.get('fullName', ''),
                'roll_number': m.get('roll_number', ''),
                'profile_picture': m.get('profile_picture'),
                'present': counts['present'],
                'absent': counts['absent'],
                'leave': counts['leave'],
                'college_work': counts['college_work'],
                'attended': attended,
                'total': total_sessions,
                'percentage': pct,
                'below_threshold': pct is not None and pct < threshold,
            })

        return jsonify({'students': result, 'total_sessions': total_sessions, 'threshold': threshold}), 200
    except Exception as e:
        logger.error(f"CR subject summary error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


@attendance_bp.route('/semester/<semester_id>/cr-roll/<session_id>/<student_id>', methods=['POST'])
@token_required
def cr_mark_student(semester_id, session_id, student_id):
    """CR marks a single student present, absent, or on leave."""
    from database import get_db
    try:
        user_id = request.user['user_id']
        db = get_db()
        semester, classroom, is_cr = _check_access(db, semester_id, user_id)
        if not semester or not is_cr:
            return jsonify({'error': 'CR access required'}), 403

        try:
            sess = db.attendance_sessions.find_one({
                '_id': ObjectId(session_id),
                'semester_id': semester_id,
                'status': 'happened',
            })
        except Exception:
            return jsonify({'error': 'Session not found'}), 404
        if not sess:
            return jsonify({'error': 'Session must be marked happened first'}), 400

        # CR can always mark any student regardless of subject mode
        try:
            if ObjectId(student_id) not in classroom.get('members', []):
                return jsonify({'error': 'Student not in classroom'}), 403
        except Exception:
            return jsonify({'error': 'Invalid student ID'}), 400

        data = request.get_json() or {}
        status = data.get('status', 'present')
        if status not in VALID_STATUSES:
            return jsonify({'error': 'Invalid status'}), 400

        now = datetime.now(timezone.utc)

        old_records = list(db.attendance_records.find({
            'semester_id': semester_id, 'student_id': student_id, 'subject': sess['subject'],
        }))
        old_N = db.attendance_sessions.count_documents({
            'semester_id': semester_id, 'subject': sess['subject'], 'status': 'happened',
        })
        old_excused = sum(1 for r in old_records if _rec_status(r) in ('leave', 'college_work'))
        old_A = sum(1 for r in old_records if _rec_status(r) == 'present')
        old_pct = round(old_A / max(old_N - old_excused, 1) * 100, 1) if old_N > 0 else 0.0

        db.attendance_records.update_one(
            {'session_id': session_id, 'student_id': student_id},
            {'$set': {
                'session_id': session_id,
                'semester_id': semester_id,
                'classroom_id': str(semester['classroom_id']),
                'student_id': student_id,
                'subject': sess['subject'],
                'date': sess['date'],
                'status': status,
                'marked_by': 'cr',
                'updated_at': now,
            }, '$setOnInsert': {'created_at': now}},
            upsert=True,
        )

        new_records = list(db.attendance_records.find({
            'semester_id': semester_id, 'student_id': student_id, 'subject': sess['subject'],
        }))
        new_excused = sum(1 for r in new_records if _rec_status(r) in ('leave', 'college_work'))
        new_A = sum(1 for r in new_records if _rec_status(r) == 'present')
        new_pct = round(new_A / max(old_N - new_excused, 1) * 100, 1) if old_N > 0 else 0.0

        try:
            from socketio_instance import socketio
            from routes.chat_routes import _connected_users
            for sid, ud in list(_connected_users.items()):
                if ud.get('user_id') == student_id:
                    socketio.emit('attendance_cr_marked', {
                        'session_id': session_id,
                        'subject': sess['subject'],
                        'date': sess['date'],
                        'slot': sess['slot'],
                        'status': status,
                        'percentage': new_pct,
                    }, to=sid)
            _check_and_notify_threshold(db, socketio, semester_id, student_id, sess['subject'], old_pct, new_pct)
        except Exception:
            pass

        return jsonify({'ok': True, 'status': status, 'percentage': new_pct}), 200
    except Exception as e:
        logger.error(f"CR mark student error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


# ── Defaulter report ──────────────────────────────────────────────────────────


@attendance_bp.route('/semester/<semester_id>/defaulters', methods=['GET'])
@token_required
def get_defaulters(semester_id):
    """
    CR fetches all students below (or within at_risk_pct of) their subject threshold.
    Query params:
      subject   — filter to one subject (optional)
      at_risk   — float, include students within this many % of threshold (default 5)
    """
    from database import get_db
    try:
        user_id = request.user['user_id']
        db = get_db()
        semester, classroom, is_cr = _check_access(db, semester_id, user_id)
        if not semester or not is_cr:
            return jsonify({'error': 'CR access required'}), 403

        subject_filter = request.args.get('subject')
        at_risk_margin = float(request.args.get('at_risk', 5.0))
        today_str = date.today().isoformat()

        settings = _get_settings(db, semester_id)
        sem_threshold = settings.get('threshold', 75.0)

        all_configs = {
            c['subject']: c
            for c in db.subject_attendance_config.find({'semester_id': semester_id})
        }

        sess_query = {
            'semester_id': semester_id,
            'status': {'$in': ['happened', 'pending']},
            'date': {'$lte': today_str},
        }
        if subject_filter:
            sess_query['subject'] = subject_filter
        happened = list(db.attendance_sessions.find(sess_query))

        # Group sessions by subject
        subj_sessions = {}
        for sess in happened:
            subj_sessions.setdefault(sess['subject'], []).append(str(sess['_id']))

        # Get all members
        members = list(db.users.find(
            {'_id': {'$in': classroom.get('members', [])}},
            {'_id': 1, 'username': 1, 'fullName': 1},
        ))

        _, cal = _get_holiday_dates(db, semester_id)
        sem_end = (cal or {}).get('semester_end') if cal else None
        future = _project_future(db, semester_id, today_str, sem_end) if sem_end else {}

        result_by_subject = {}

        for subj, sess_ids in subj_sessions.items():
            cfg = all_configs.get(subj, {})
            mode = cfg.get('tracking_mode', 'self') if cfg else _subject_mode(db, semester_id, subj, settings)
            if mode == 'off':
                continue
            threshold = cfg.get('required_pct') if cfg and cfg.get('required_pct') is not None else sem_threshold
            total = len(sess_ids)
            R = future.get(subj, 0)

            defaulters = []
            at_risk = []

            for m in members:
                mid = str(m['_id'])
                recs = list(db.attendance_records.find({
                    'student_id': mid,
                    'session_id': {'$in': sess_ids},
                }))
                rec_map = {r['session_id']: r for r in recs}
                attended = sum(1 for sid in sess_ids if _rec_status(rec_map.get(sid)) == 'present')
                leaves_count = sum(1 for sid in sess_ids if _rec_status(rec_map.get(sid)) == 'leave')
                cw_count = sum(1 for sid in sess_ids if _rec_status(rec_map.get(sid)) == 'college_work')
                excused = leaves_count + cw_count

                pct, leaves_left, must_attend, recoverable = _calc_subject_stats(
                    attended, total, excused, R, threshold
                )

                entry = {
                    'user_id': mid,
                    'username': m.get('username', ''),
                    'full_name': m.get('fullName', ''),
                    'attended': attended,
                    'total': total,
                    'leaves_count': leaves_count,
                    'cw_count': cw_count,
                    'percentage': pct,
                    'must_attend': must_attend,
                    'leaves_left': leaves_left,
                    'recoverable': recoverable,
                }

                if pct < threshold:
                    defaulters.append(entry)
                elif pct < threshold + at_risk_margin:
                    at_risk.append(entry)

            if defaulters or at_risk:
                result_by_subject[subj] = {
                    'threshold': threshold,
                    'total_sessions': total,
                    'defaulters': sorted(defaulters, key=lambda x: x['percentage']),
                    'at_risk': sorted(at_risk, key=lambda x: x['percentage']),
                }

        return jsonify({'report': result_by_subject, 'at_risk_margin': at_risk_margin}), 200
    except Exception as e:
        logger.error(f"Get defaulters error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


# ── Excel export ──────────────────────────────────────────────────────────────


def _build_excel(db, semester_id, classroom, settings, subjects_filter=None):
    """
    Build an Excel workbook with one sheet per subject.
    Columns: Roll/Username | Name | <date slot> ... | Total | Present | Leave | Absent | %
    Returns BytesIO.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return None, 'openpyxl not installed. Run: pip install openpyxl'

    today_str = date.today().isoformat()
    sem_threshold = settings.get('threshold', 75.0)
    all_configs = {
        c['subject']: c
        for c in db.subject_attendance_config.find({'semester_id': semester_id})
    }

    sess_query = {'semester_id': semester_id, 'status': {'$in': ['happened', 'pending']}, 'date': {'$lte': today_str}}
    if subjects_filter:
        sess_query['subject'] = {'$in': subjects_filter}
    happened = list(db.attendance_sessions.find(sess_query).sort([('date', 1), ('slot', 1)]))

    subj_sessions = {}
    for sess in happened:
        subj_sessions.setdefault(sess['subject'], []).append(sess)

    members = list(db.users.find(
        {'_id': {'$in': classroom.get('members', [])}},
        {'_id': 1, 'username': 1, 'fullName': 1},
    ).sort('fullName', 1))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    header_fill = PatternFill('solid', fgColor='4F81BD')
    header_font = Font(bold=True, color='FFFFFF')
    red_fill = PatternFill('solid', fgColor='FFC7CE')
    orange_fill = PatternFill('solid', fgColor='FFEB9C')
    green_fill = PatternFill('solid', fgColor='C6EFCE')

    for subj, sessions in subj_sessions.items():
        cfg = all_configs.get(subj, {})
        threshold = cfg.get('required_pct') if cfg and cfg.get('required_pct') is not None else sem_threshold

        ws = wb.create_sheet(title=subj[:31])  # sheet name limit

        # Header row
        headers = ['Roll No', 'Name']
        for sess in sessions:
            headers.append(f"{sess['date']}\n{sess['slot']}")
        headers += ['Total', 'Present', 'Leave', 'College Work', 'Absent', '%']

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, horizontal='center')

        sess_ids = [str(s['_id']) for s in sessions]
        all_records = list(db.attendance_records.find({
            'session_id': {'$in': sess_ids},
            'classroom_id': str(classroom['_id']),
        }))
        rec_map = {}
        for r in all_records:
            rec_map.setdefault(r['student_id'], {})[r['session_id']] = r

        for row_idx, m in enumerate(members, 2):
            mid = str(m['_id'])
            student_recs = rec_map.get(mid, {})
            ws.cell(row=row_idx, column=1, value=m.get('username', mid))
            ws.cell(row=row_idx, column=2, value=m.get('fullName', ''))

            present_count = leave_count = cw_count = absent_count = 0
            for col_idx, sess in enumerate(sessions, 3):
                sid = str(sess['_id'])
                rec = student_recs.get(sid)
                status = _rec_status(rec)
                if status == 'present':
                    label, present_count = 'P', present_count + 1
                elif status == 'leave':
                    label, leave_count = 'L', leave_count + 1
                elif status == 'college_work':
                    label, cw_count = 'CW', cw_count + 1
                elif status == 'absent':
                    label, absent_count = 'A', absent_count + 1
                else:
                    label = '—'
                    absent_count += 1  # unmarked counts as absent in report
                c = ws.cell(row=row_idx, column=col_idx, value=label)
                c.alignment = Alignment(horizontal='center')

            total = len(sessions)
            excused = leave_count + cw_count
            effective = total - excused
            pct = round(present_count / effective * 100, 1) if effective > 0 else 0.0

            ws.cell(row=row_idx, column=len(headers) - 5, value=total)
            ws.cell(row=row_idx, column=len(headers) - 4, value=present_count)
            ws.cell(row=row_idx, column=len(headers) - 3, value=leave_count)
            ws.cell(row=row_idx, column=len(headers) - 2, value=cw_count)
            ws.cell(row=row_idx, column=len(headers) - 1, value=absent_count)
            pct_cell = ws.cell(row=row_idx, column=len(headers), value=f'{pct}%')
            pct_cell.alignment = Alignment(horizontal='center')
            if pct < threshold:
                pct_cell.fill = red_fill
            elif pct < threshold + 5:
                pct_cell.fill = orange_fill
            else:
                pct_cell.fill = green_fill

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 24
        for i in range(3, len(headers) - 3):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, None


def _auth_from_request():
    """Read and decode JWT from Authorization header or ?token= query param. Returns user dict or None."""
    import jwt as pyjwt
    SECRET_KEY = os.getenv('JWT_SECRET', 'dev-secret-change-in-production')
    token = request.args.get('token') or request.headers.get('Authorization', '')
    if token.startswith('Bearer '):
        token = token[7:]
    if not token:
        return None
    try:
        return pyjwt.decode(token, SECRET_KEY, algorithms=['HS256'])
    except pyjwt.PyJWTError:
        return None


@attendance_bp.route('/semester/<semester_id>/subject/<path:subject>/export/excel', methods=['GET'])
def export_subject_excel(semester_id, subject):
    """Download Excel sheet for one subject. Also marks last_exported_at. Accepts ?token= for browser downloads."""
    from database import get_db
    try:
        user_data = _auth_from_request()
        if not user_data:
            return jsonify({'error': 'Token is missing'}), 401
        user_id = user_data['user_id']
        db = get_db()
        semester, classroom, is_cr = _check_access(db, semester_id, user_id)
        if not semester or not is_cr:
            return jsonify({'error': 'CR access required'}), 403

        settings = _get_settings(db, semester_id)
        buf, err = _build_excel(db, semester_id, classroom, settings, subjects_filter=[subject])
        if err:
            return jsonify({'error': err}), 500

        # Mark exported
        now = datetime.now(timezone.utc)
        db.subject_attendance_config.update_one(
            {'semester_id': semester_id, 'subject': subject},
            {'$set': {'last_exported_at': now.isoformat(), 'updated_at': now},
             '$setOnInsert': {'created_at': now, 'archived_periods': [], 'tracking_mode': 'self'}},
            upsert=True,
        )

        filename = f"attendance_{subject}_{date.today().isoformat()}.xlsx"
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        logger.error(f"Export subject excel error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


@attendance_bp.route('/semester/<semester_id>/export/excel', methods=['GET'])
def export_all_excel(semester_id):
    """Download Excel workbook for all subjects (one sheet each). Accepts ?token= for browser downloads."""
    from database import get_db
    try:
        user_data = _auth_from_request()
        if not user_data:
            return jsonify({'error': 'Token is missing'}), 401
        user_id = user_data['user_id']
        db = get_db()
        semester, classroom, is_cr = _check_access(db, semester_id, user_id)
        if not semester or not is_cr:
            return jsonify({'error': 'CR access required'}), 403

        settings = _get_settings(db, semester_id)
        buf, err = _build_excel(db, semester_id, classroom, settings)
        if err:
            return jsonify({'error': err}), 500

        now = datetime.now(timezone.utc)
        # Mark all CR-mode subjects as exported
        for cfg in db.subject_attendance_config.find({
            'semester_id': semester_id, 'tracking_mode': 'cr',
        }):
            db.subject_attendance_config.update_one(
                {'_id': cfg['_id']},
                {'$set': {'last_exported_at': now.isoformat(), 'updated_at': now}},
            )

        filename = f"attendance_{semester_id}_{date.today().isoformat()}.xlsx"
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        logger.error(f"Export all excel error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


# ── Defaulter report export ───────────────────────────────────────────────────


@attendance_bp.route('/semester/<semester_id>/defaulters/export/excel', methods=['GET'])
def export_defaulters_excel(semester_id):
    """Download defaulter report as Excel. Accepts ?token= for browser downloads."""
    from database import get_db
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return jsonify({'error': 'openpyxl not installed'}), 500

    try:
        user_data = _auth_from_request()
        if not user_data:
            return jsonify({'error': 'Token is missing'}), 401
        user_id = user_data['user_id']
        db = get_db()
        semester, classroom, is_cr = _check_access(db, semester_id, user_id)
        if not semester or not is_cr:
            return jsonify({'error': 'CR access required'}), 403

        at_risk_margin = float(request.args.get('at_risk', 5.0))
        today_str = date.today().isoformat()
        settings = _get_settings(db, semester_id)
        sem_threshold = settings.get('threshold', 75.0)

        all_configs = {c['subject']: c for c in db.subject_attendance_config.find({'semester_id': semester_id})}
        happened = list(db.attendance_sessions.find({
            'semester_id': semester_id,
            'status': {'$in': ['happened', 'pending']},
            'date': {'$lte': today_str},
        }))
        subj_sessions = {}
        for sess in happened:
            subj_sessions.setdefault(sess['subject'], []).append(str(sess['_id']))

        members = list(db.users.find(
            {'_id': {'$in': classroom.get('members', [])}},
            {'_id': 1, 'username': 1, 'fullName': 1},
        ).sort('fullName', 1))

        _, cal = _get_holiday_dates(db, semester_id)
        sem_end = (cal or {}).get('semester_end') if cal else None
        future = _project_future(db, semester_id, today_str, sem_end) if sem_end else {}

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill('solid', fgColor='4F81BD')
        header_font = Font(bold=True, color='FFFFFF')
        red_fill    = PatternFill('solid', fgColor='FFC7CE')
        orange_fill = PatternFill('solid', fgColor='FFEB9C')

        for subj, sess_ids in sorted(subj_sessions.items()):
            cfg = all_configs.get(subj, {})
            mode = cfg.get('tracking_mode', 'self') if cfg else _subject_mode(db, semester_id, subj, settings)
            if mode == 'off':
                continue
            threshold = cfg.get('required_pct') if cfg and cfg.get('required_pct') is not None else sem_threshold
            total = len(sess_ids)
            R = future.get(subj, 0)

            rows = []
            for m in members:
                mid = str(m['_id'])
                recs = list(db.attendance_records.find({'student_id': mid, 'session_id': {'$in': sess_ids}}))
                rec_map = {r['session_id']: r for r in recs}
                attended = sum(1 for sid in sess_ids if _rec_status(rec_map.get(sid)) == 'present')
                excused  = sum(1 for sid in sess_ids if _rec_status(rec_map.get(sid)) in ('leave', 'college_work'))
                pct, _, must_attend, recoverable = _calc_subject_stats(attended, total, excused, R, threshold)
                if pct >= threshold + at_risk_margin:
                    continue
                rows.append({
                    'username': m.get('username', ''),
                    'full_name': m.get('fullName', ''),
                    'attended': attended,
                    'excused': excused,
                    'total': total,
                    'pct': pct,
                    'must_attend': must_attend,
                    'recoverable': recoverable,
                    'defaulter': pct < threshold,
                })

            if not rows:
                continue

            ws = wb.create_sheet(title=subj[:31])
            headers = ['Roll No', 'Name', 'Attended', 'Excused', 'Total', '%', 'Must Attend', 'Status']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for ri, row in enumerate(sorted(rows, key=lambda x: x['pct']), 2):
                ws.cell(row=ri, column=1, value=row['username'])
                ws.cell(row=ri, column=2, value=row['full_name'])
                ws.cell(row=ri, column=3, value=row['attended'])
                ws.cell(row=ri, column=4, value=row['excused'])
                ws.cell(row=ri, column=5, value=row['total'])
                pct_cell = ws.cell(row=ri, column=6, value=f"{row['pct']}%")
                pct_cell.alignment = Alignment(horizontal='center')
                pct_cell.fill = red_fill if row['defaulter'] else orange_fill
                ws.cell(row=ri, column=7, value=row['must_attend'] if row['defaulter'] else 0)
                ws.cell(row=ri, column=8, value='Defaulter' if row['defaulter'] else 'At Risk')

            ws.column_dimensions['A'].width = 14
            ws.column_dimensions['B'].width = 24
            for col in range(3, 9):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12

        if not wb.sheetnames:
            ws = wb.create_sheet('No Defaulters')
            ws.cell(row=1, column=1, value='No defaulters or at-risk students found.')

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"defaulters_{semester_id}_{date.today().isoformat()}.xlsx"
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        logger.error(f"Export defaulters excel error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


# ── Attendance proof attachments ──────────────────────────────────────────────


@attendance_bp.route('/semester/<semester_id>/record/<session_id>/attachment', methods=['POST'])
@token_required
def upload_attachment(semester_id, session_id):
    """Student uploads a proof file (for leave or college_work records)."""
    from database import get_db
    try:
        user_id = request.user['user_id']
        db = get_db()
        semester, _, _ = _check_access(db, semester_id, user_id)
        if not semester:
            return jsonify({'error': 'Access denied'}), 403

        rec = db.attendance_records.find_one({'session_id': session_id, 'student_id': user_id})
        if not rec:
            return jsonify({'error': 'No attendance record found — mark the session first'}), 404
        if _rec_status(rec) not in ('leave', 'college_work'):
            return jsonify({'error': 'Attachments only allowed on leave or college work records'}), 400

        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        f = request.files['file']
        if not f.filename:
            return jsonify({'error': 'No file selected'}), 400

        os.makedirs(PROOF_DIR, exist_ok=True)
        ext = os.path.splitext(f.filename)[1].lower()
        stored_name = f"{uuid.uuid4().hex}{ext}"
        f.save(os.path.join(PROOF_DIR, stored_name))

        # Remove old attachment file if present
        old_att = rec.get('attachment')
        if old_att and old_att.get('stored_name'):
            try:
                os.remove(os.path.join(PROOF_DIR, old_att['stored_name']))
            except OSError:
                pass

        now = datetime.now(timezone.utc)
        db.attendance_records.update_one(
            {'_id': rec['_id']},
            {'$set': {'attachment': {'original_name': f.filename, 'stored_name': stored_name}, 'updated_at': now}},
        )

        return jsonify({'ok': True, 'attachment': {'original_name': f.filename, 'stored_name': stored_name}}), 200
    except Exception as e:
        logger.error(f"Upload attachment error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


@attendance_bp.route('/semester/<semester_id>/record/<session_id>/attachment', methods=['DELETE'])
@token_required
def delete_attachment(semester_id, session_id):
    """Remove a proof attachment from a leave/college_work record."""
    from database import get_db
    try:
        user_id = request.user['user_id']
        db = get_db()
        semester, _, _ = _check_access(db, semester_id, user_id)
        if not semester:
            return jsonify({'error': 'Access denied'}), 403

        rec = db.attendance_records.find_one({'session_id': session_id, 'student_id': user_id})
        if not rec or not rec.get('attachment'):
            return jsonify({'error': 'No attachment found'}), 404

        stored_name = rec['attachment'].get('stored_name')
        if stored_name:
            try:
                os.remove(os.path.join(PROOF_DIR, stored_name))
            except OSError:
                pass

        db.attendance_records.update_one(
            {'_id': rec['_id']},
            {'$unset': {'attachment': ''}, '$set': {'updated_at': datetime.now(timezone.utc)}},
        )
        return jsonify({'ok': True}), 200
    except Exception as e:
        logger.error(f"Delete attachment error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


@attendance_bp.route('/my-proofs', methods=['GET'])
@token_required
def list_my_proofs(semester_id=None):
    """Return all proof attachments the current user has uploaded (across all semesters)."""
    from database import get_db
    try:
        user_id = request.user['user_id']
        db = get_db()

        recs = list(db.attendance_records.find({
            'student_id': user_id,
            'attachment': {'$exists': True, '$ne': None},
        }))

        result = []
        for rec in recs:
            att = rec.get('attachment')
            if not att or not att.get('stored_name'):
                continue
            result.append({
                'session_id': rec.get('session_id'),
                'semester_id': rec.get('semester_id'),
                'subject': rec.get('subject'),
                'date': rec.get('date'),
                'status': _rec_status(rec),
                'original_name': att.get('original_name'),
                'stored_name': att.get('stored_name'),
            })

        result.sort(key=lambda x: x.get('date') or '', reverse=True)
        return jsonify({'proofs': result}), 200
    except Exception as e:
        logger.error(f"List my proofs error: {e}")
        return jsonify({'error': 'Internal server error'}), 500


@attendance_bp.route('/proof/<path:filename>', methods=['GET'])
def serve_proof(filename):
    """Serve an attendance proof file. Accepts token via ?token= query param or Authorization header."""
    import jwt as pyjwt
    SECRET_KEY = os.getenv('JWT_SECRET', 'dev-secret-change-in-production')
    token = request.args.get('token') or request.headers.get('Authorization', '')
    if token.startswith('Bearer '):
        token = token[7:]
    if not token:
        return jsonify({'error': 'Token required'}), 401
    try:
        pyjwt.decode(token, SECRET_KEY, algorithms=['HS256'])
    except pyjwt.PyJWTError:
        return jsonify({'error': 'Invalid token'}), 401
    return send_from_directory(PROOF_DIR, filename)
